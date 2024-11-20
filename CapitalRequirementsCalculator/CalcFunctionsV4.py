import pandas as pd
import xlsxwriter
from xlsxwriter.utility import xl_rowcol_to_cell #temporary
import openpyxl
from openpyxl.chart import PieChart, Reference
import xlsxwriter
from tkinter import *
import tkinter as tk
import os
import glob
import re

class SFTButton(Button):
    def __init__(self,master,**config):
        super(SFTButton, self).__init__(master,**config)
        self.config(font='Calibri', fg='white', bg='#003B70', width=30)
        
def goto_page(from_page, to_page):
    from_page.grid_remove()
    to_page.grid(row=0, column=0)

#***** https://urlzs.com/k6aj1 Adapted to use my files
def open_report(file_name):
    os.startfile(file_name)
#*****

def final_transaction(row, exposure_sheet): 
    if exposure_sheet.iloc[row, 6] == 'B': #Row starts at 0
        return 'Repurchase Transaction'
    if exposure_sheet.iloc[row, 6] == 'LTRR':
        return 'Repurchase Transaction'
    if exposure_sheet.iloc[row, 6] == 'REG':
        return 'Repurchase Transaction'
    if exposure_sheet.iloc[row, 6] == 'REGR':
        return 'Repurchase Transaction'
    if exposure_sheet.iloc[row, 6] == 'REV':
        return 'Repurchase Transaction'
    if exposure_sheet.iloc[row, 6] == 'REVR':
        return 'Repurchase Transaction'
    if exposure_sheet.iloc[row, 6] == 'S':
        return 'Repurchase Transaction'
    if exposure_sheet.iloc[row, 6] == 'BVC':
        return 'Securities or commodities'
    if exposure_sheet.iloc[row, 6] == 'BVDV':
        return 'Securities or commodities'
    if exposure_sheet.iloc[row, 6] == 'BVP':
        return 'Securities or commodities'
    if exposure_sheet.iloc[row, 6] == 'LVC':
        return 'Securities or commodities'
    if exposure_sheet.iloc[row, 6] == 'LVDV':
        return 'Securities or commodities'
    if exposure_sheet.iloc[row, 6] == 'LVP':
        return 'Securities or commodities'
    if exposure_sheet.iloc[row, 6] == 'LONG':
        return 'Margin lending transcation'
    if exposure_sheet.iloc[row, 6] == 'SHORT':
        return 'Margin lending transcation'
    if pd.isna(exposure_sheet.iloc[row, 6]) == True:
        return 'Error'

def collateral_type(row, optima_sheet):
    return optima_sheet.iloc[row+1, 6]

def cash_security(row, exposure_sheet):
    security_1 = 0 #1 security for each combination that makes security
    security_2 = 0
    security_3 = 0
    if exposure_sheet.iloc[row, 11] == 'Exposure':  #'Or' and 'and' not compatible with pandas
        security_1 += 1
    if exposure_sheet.iloc[row, 12] == 'S':
        security_1 += 1
    if exposure_sheet.iloc[row, 13] == 'S':
        security_1 += 1
    if exposure_sheet.iloc[row, 11] == 'Exposure':
        security_2 += 1
    if exposure_sheet.iloc[row, 12] == 'B':
        security_2 +=1
    if exposure_sheet.iloc[row, 13] == 'CE':
        security_2 += 1
    if exposure_sheet.iloc[row, 11] == 'Collateral':
        security_3 += 1
    if exposure_sheet.iloc[row, 12] == 'B':
        security_3 += 1
    if exposure_sheet.iloc[row, 13] == 'S':
        security_3 += 1
    if security_1 > 2:
        return 'Security'
    elif security_2 > 2:
        return 'Security'
    elif security_3 > 2:
        return 'Security'
    else:
        return 'Cash'


def master_netting(row, exposure_sheet):
    if exposure_sheet.iloc[row, 16] > 0:
        return 'YES'
    else:
        return 'NO'
        

def legal_enforceability(row, exposure_sheet): #AND SINGLE NET, VALUE STABILITY, RIGHT TO LIQUIDATE
    if exposure_sheet.iloc[row, 19] == 'Y':
        return 'YES'

def liquid_illiquid(row, EE_sheet):
    return EE_sheet.iloc[row, 1]

'''def sufficiently_liquid(row, exposure_sheet):
    yes = 0
    if exposure_sheet.iloc[row, 16] == 0:
        yes +=1
    if exposure_sheet.iloc[row, 22] == 'N':
        yes += 1
    if yes > 1:
        return 'YES'
    else:
        return 'YES''' #to be finished MNA 

def correlated_collateral(row, EE_sheet):
    return EE_sheet.iloc[row, 2]

def correlation_acceptability(row, EE_sheet):
    if correlated_collateral(row, EE_sheet) == 'N':
        return 'YES'
    else:
        return 'NO'

def SWWR_flag(row, optima_sheet):
    return optima_sheet.iloc[row+1, 2]

def GFP_ID(row, exposure_sheet):
    return exposure_sheet.iloc[row, 3]

def haircut_security(row, exposure_sheet, optima_sheet):
    if exposure_sheet.iloc[row, 13] == 'CE':
        return 'CE'
    else:
        return optima_sheet.iloc[row+1, 10]

def basel_asset(row, exposure_sheet, optima_sheet):
    if exposure_sheet.iloc[row, 13] == 'CE':
        return 'OTH'
    else:
        return optima_sheet.iloc[row+1, 3]

def currency_exposure(row, optima_sheet):
    if str(optima_sheet.iloc[row+1, 9]) == 'nan':
        return '#N/A'
    else:
        return optima_sheet.iloc[row+1, 9]

def value_collateral(row, exposure_sheet):
    if exposure_sheet.iloc[row, 11] == 'Collateral':
        value = abs(int(exposure_sheet.iloc[row, 10]))
    else:
        value = abs(int(exposure_sheet.iloc[row, 9]))
    return value

def general_requirements(row, exposure_sheet):
    no = 0
    if exposure_sheet.iloc[row, 20] == 'NO':
        no +=1
    elif exposure_sheet.iloc[row, 21] == 'NO':
        no +=1
    elif exposure_sheet.iloc[row, 24] == 'NO':
        no += 1
    elif exposure_sheet.iloc[row, 25] == 'NO':
        no += 1
    if no > 0:
        return 'NO'
    else:
        return 'YES'

def collateral_cat(row, exposure_sheet):
    concat = str(exposure_sheet.iloc[row, 31])+str(exposure_sheet.iloc[row, 45])+str(exposure_sheet.iloc[row, 80]) 
    return concat

def eligible_collateral(row, exposure_sheet, lists_sheet, rows_lists):  #No
    col_type = exposure_sheet.iloc[row, 73]
    for rows in range(rows_lists):
        if lists_sheet.iloc[rows, 7] == col_type:
            eligible = exposure_sheet.iloc[row, 8]
            return eligible
        else:
            return 'No'

def credit_maturity(row, exposure_sheet):
    return exposure_sheet.iloc[row, 66]
    
def long_ratingsSP(row, exposure_sheet):
    rating = exposure_sheet.iloc[row, 33]
    if pd.isna(rating) == False:
        CQS_rating = working_ratings[rating]
        return CQS_rating
    else:
        return 0

def long_ratingsM(row, exposure_sheet): 
    rating = exposure_sheet.iloc[row, 34]
    if pd.isna(rating) == False:
        CQS_rating = working_ratings[rating]
        return CQS_rating
    else:
        return 0

def long_ratingsF(row, exposure_sheet):
    rating = exposure_sheet.iloc[row, 35]
    if pd.isna(rating) == False:
        CQS_rating = working_ratings[rating]
        return CQS_rating
    else:
        return 0

def long_ratingsDBRS(row, exposure_sheet):
    rating = exposure_sheet.iloc[row, 36]
    if pd.isna(rating) == False:
        CQS_rating = working_ratings[rating]
        return CQS_rating
    else:
        return 0

def short_ratingsSP(row, exposure_sheet):
    rating = exposure_sheet.iloc[row, 37]
    if pd.isna(rating) == False:
        CQS_rating = working_ratings[rating]
        return CQS_rating
    else:
        return 0

def short_ratingsM(row, exposure_sheet):
    rating = exposure_sheet.iloc[row, 38]
    if pd.isna(rating) == False:
        CQS_rating = working_ratings[rating]
        return CQS_rating
    else:
        return 0

def short_ratingsF(row, exposure_sheet):
    rating = exposure_sheet.iloc[row, 39]
    if pd.isna(rating) == False:
        CQS_rating = working_ratings[rating]
        return CQS_rating
    else:
        return 0

def short_ratingsDBRS(row, exposure_sheet):
    rating = exposure_sheet.iloc[row, 40]
    if pd.isna(rating) == False:
        CQS_rating = working_ratings[rating]
        return CQS_rating
    else:
        return 0

long_countif = []
def long_ratingCount(exposure_sheet, rows_exposure):
    for row in range(rows_exposure):
        col_filled = 0 
        for column in range(33, 37, 1):
            if pd.isna(exposure_sheet.iloc[row, column]) == False and exposure_sheet.iloc[row, column] != 0:
                col_filled +=1
        long_countif.append(col_filled)

def longTermCQS(row, exposure_sheet):
    longterm_list = []
    if exposure_sheet.iloc[row, 79] == 1:
        for column in range(75, 78):
            if exposure_sheet.iloc[row, column] > 0:
                longterm_val = exposure_sheet.iloc[row, column]
    elif exposure_sheet.iloc[row, 79] >= 2:
        for column in range(75, 78):
            if exposure_sheet.iloc[row, column] > 0:
                longterm_list.append(exposure_sheet.iloc[row, column])
        longterm_list.sort()
        longterm_val = longterm_list[1]
    else:
        longterm_val = 0
    return longterm_val

short_countif = []
def short_ratingCount(exposure_sheet, rows_exposure):
    for row in range(rows_exposure):
        col_filled = 0 
        for column in range(38, 41, 1):
            if pd.isna(exposure_sheet.iloc[row, column]) == False:
                col_filled +=1
        short_countif.append(col_filled)

def shortTermCQS(row, exposure_sheet):
    shortterm_list = []
    if exposure_sheet.iloc[row, 79] == 1:
        for column in range(75, 78):
            if exposure_sheet.iloc[row, column] > 0:
                shortterm_val = exposure_sheet.iloc[row, column]
    elif exposure_sheet.iloc[row, 79] >= 2:
        for column in range(75, 78):
            if exposure_sheet.iloc[row, column] > 0:
                shortterm_list.append(exposure_sheet.iloc[row, column])
        shortterm_list.sort()
        shortterm_val = shortterm_list[1]
    else:
        shortterm_val = 0
    return shortterm_val

def recognised_exchange(row, exposure_sheet, lists_sheet):
    valid_exchanges = ['EURONEXT PARIS',
                       'BOERSE BERLIN (REGULIERTER MARKT)',
                       'BOERSE BERLIN (BERLIN SECOND REGULATED MARKET)',
                       'BOERSE DUESSELDORF (REGULIERTER MARKT)',
                       'BOERSE DUESSELDORF - QUOTRIX (REGULIERTER MARKT)',
                       'BOERSE BERLIN EQUIDUCT TRADING (REGULIERTER MARKT)'
                       ]
    if exposure_sheet.iloc[row, 42] in valid_exchanges:
        return 'YES'
    else:
        return 'NO'

def basic_liquidation(row, exposure_sheet):
    buy_ind = exposure_sheet.iloc[row, 6]
    liquid_period = working_basic[buy_ind]
    return liquid_period

def num_trades(row, exposure_sheet):
    if exposure_sheet.iloc[row, 7] == 123:
        return 9
    elif exposure_sheet.iloc[row, 7] == 456:
        return 11

def large_netting(row, exposure_sheet):
    if exposure_sheet.iloc[row, 92] > 5000:
        return 20
    else:
        return exposure_sheet.iloc[row, 91]

def adjusted_margin(row, exposure_sheet):
    if exposure_sheet.iloc[row, 94] > 2:
        return int(exposure_sheet.iloc[row, 93])*2
    else:
        return int(exposure_sheet.iloc[row, 93])

def residual_maturity(row, exposure_sheet):
    if exposure_sheet.iloc[row, 58] == 0:
        return ''
    elif exposure_sheet.iloc[row, 58] < 366:
        return '<1YR'
    elif exposure_sheet.iloc[row, 58] < 1827:
        return '1-5YR'
    else:
        return '>5YR'

def vol_adj(row, exposure_sheet):
    collateral_type = exposure_sheet.iloc[row, 73]
    res_maturity = exposure_sheet.iloc[row, 96]
    vol = collateral_type+res_maturity
    return vol

def ten_voladj(row, exposure_sheet, volatility_sheet, rows_volatility):
    vol = exposure_sheet.iloc[row, 97]
    for i in range(rows_volatility):
        if vol == volatility_sheet.iloc[i, 0]:
            return volatility_sheet.iloc[i, 1]

def security_volatility(row, exposure_sheet):
    if exposure_sheet.iloc[row, 48] == 'YES' and exposure_sheet.iloc(row, 127) == exposure_sheet.iloc[row, 56]:
        return exposure_sheet.iloc[row, 130]
    else:
        adj_margin = exposure_sheet.iloc[row, 95]/10
        ten_vol = exposure_sheet.iloc[row, 98]
        result = (adj_margin**0.5)*ten_vol
        #***https://stackoverflow.com/questions/28142688/how-to-turn-input-number-into-a-percentage-in-python
        return "{:.1%}".format(result/100)
        #****
        
def security_exposure(row, exposure_sheet):
    if exposure_sheet.iloc[row, 11] == 'Exposure':
        txn_e = exposure_sheet.iloc[row, 9]
        security_vol = exposure_sheet.iloc[row,99]
        return txn_e*(1+float(security_vol[0:3]))
    else:
        txn_c = exposure_sheet.iloc[row, 10]
        security_vol = exposure_sheet.iloc[row,99]
        return txn_c*(1-float(security_vol[0:3]))

def fx_mismatch(row, exposure_sheet):
    if exposure_sheet.iloc[row, 56] == exposure_sheet.iloc[row, 57]:
        return 'No'
    else:
        return 'Yes'

def tenday_fx(row, exposure_sheet):
    if exposure_sheet.iloc[row, 101] == 'Yes':
        return "{:.1%}".format(0.08)
    else:
        return ("{:.1%}".format(0.00))

def security_fx(row, exposure_sheet):
    txn_e = exposure_sheet.iloc[row, 9]
    txn_c = exposure_sheet.iloc[row, 10]
    security_vol = exposure_sheet.iloc[row, 99]
    ten_fx = exposure_sheet.iloc[row, 102]
    if exposure_sheet.iloc[row, 11] == 'Exposure':
        return txn_e*(1+float(security_vol[0:3])+float(ten_fx[0:3]))
    else:
        return txn_c*(1-float(security_vol[0:3])-float(ten_fx[0:3]))
        
def sov_SP(row, exposure_sheet):
    rating = exposure_sheet.iloc[row, 107]
    sov_rating = working_ratings[rating]
    return sov_rating

def sov_M(row, exposure_sheet):
    rating = exposure_sheet.iloc[row, 108]
    sov_rating = working_ratings[rating]
    return sov_rating

def sov_F(row, exposure_sheet):
    rating = exposure_sheet.iloc[row, 109]
    sov_rating = working_ratings[rating]
    return sov_rating

def sov_DBRS(row, exposure_sheet):
    rating = exposure_sheet.iloc[row, 110]
    sov_rating = working_ratings[rating]
    return sov_rating

sov_countif = []
def sov_ratingCount(exposure_sheet, rows_exposure):
    for row in range(rows_exposure):
        col_filled = 0 
        for column in range(111, 114, 1):
            if pd.isna(exposure_sheet.iloc[row, column]) == False and exposure_sheet.iloc[row, column] != 0:
                col_filled +=1
        sov_countif.append(col_filled)

def sov_CQS(row, exposure_sheet):
    sovdom_list = []
    if exposure_sheet.iloc[row, 115] == 1: #Only 1 value in sov ratings
        for column in range(111, 114):
            if exposure_sheet.iloc[row, column] > 0:
                sovdom_val = exposure_sheet.iloc[row, column]
    elif exposure_sheet.iloc[row, 79] >= 2:
        for column in range(111, 114):
            if exposure_sheet.iloc[row, column] > 0: 
                sovdom_list.append(exposure_sheet.iloc[row, column])
        sovdom_list.sort()
        sovdom_val = sovdom_list[1]
    else:
        sovdom_val = 0
    return sovdom_val

def sovF_SP(row, exposure_sheet):
    rating = exposure_sheet.iloc[row, 117]
    sov_rating = working_ratings[rating]
    return sov_rating

def sovF_M(row, exposure_sheet):
    rating = exposure_sheet.iloc[row, 118]
    sov_rating = working_ratings[rating]
    return sov_rating

def sovF_F(row, exposure_sheet):
    rating = exposure_sheet.iloc[row, 119]
    sov_rating = working_ratings[rating]
    return sov_rating

def sovF_DBRS(row, exposure_sheet):
    rating = exposure_sheet.iloc[row, 120]
    sov_rating = working_ratings[rating]
    return sov_rating

sovF_countif = []
def sovF_ratingCount(exposure_sheet, rows_exposure):
    for row in range(rows_exposure):
        col_filled = 0 
        for column in range(121, 125, 1):
            if pd.isna(exposure_sheet.iloc[row, column]) == False and exposure_sheet.iloc[row, column] != 0:
                col_filled +=1
        sovF_countif.append(col_filled)

def sovF_CQS(row, exposure_sheet):
    sovdom_list = []
    if exposure_sheet.iloc[row, 125] == 1: #Only 1 value in sov ratings
        for column in range(121, 124):
            if exposure_sheet.iloc[row, column] > 0:
                sovdom_val = exposure_sheet.iloc[row, column]
    elif exposure_sheet.iloc[row, 79] >= 2:
        for column in range(121, 124):
            if exposure_sheet.iloc[row, column] > 0: 
                sovdom_list.append(exposure_sheet.iloc[row, column])
        sovdom_list.sort()
        sovdom_val = sovdom_list[1]
    else:
        sovdom_val = 0
    return sovdom_val

def domestic_currency(row, exposure_sheet):
    currency_lookup = {'GB':'GBP', 'US':'USD', 'HK':'HKD', 'DE':'EUR', 'FR':'EUR'}
    country = exposure_sheet.iloc[row, 46]
    if country in currency_lookup:
        currency = currency_lookup[country]
        return currency
    else:
        return 'Not valid'

def sovereign_ref(row, exposure_sheet):
    ref = 'DEBT/DEBTSOV'+str(exposure_sheet.iloc[row, 126])+str(exposure_sheet.iloc[row, 96])
    return ref

def sov_volA(row, exposure_sheet, volatility_sheet, rows_volatility):
    sov_ref = exposure_sheet.iloc[row, 128]
    for i in range(rows_volatility):
        if sov_ref == volatility_sheet.iloc[i, 0]:
            return volatility_sheet.iloc[i, 1]

def sov_security(row, exposure_sheet):
    margin_disp = exposure_sheet.iloc[row, 95]
    sovVA = exposure_sheet.iloc[row, 129]
    value = ((margin_disp/10)**0.5)*sovVA
    return "{:.1%}".format(value)
    
SC_Trans = []
ML_Trans = []
RE_Trans = []    
def pie_figures(row, exposure_sheet):
    if exposure_sheet.iloc[row, 8] == 'Securities or commodities':
        num1 = int(exposure_sheet.iloc[row, 9])
        SC_Trans.append(num1)
    elif exposure_sheet.iloc[row, 8] == 'Repurchase Transaction':
        num2 = int(exposure_sheet.iloc[row, 9])
        RE_Trans.append(num2)
    elif exposure_sheet.iloc[row, 8] == 'Margin lending transcation':
        num3 = int(exposure_sheet.iloc[row, 9])
        ML_Trans.append(num3)

def pie_chart(pie_sheet):
    SC_num = sum(SC_Trans)
    ML_num = sum(ML_Trans)
    RE_num = sum(RE_Trans)
    #******* https://bit.ly/2ohaX62
    data = [['Transaction', 'Value'],['Margin Lending Transaction', ML_num], ['Repurchase Transaction', RE_num], ['Securities or commodities', SC_num]]
    for sect in data: 
        pie_sheet.append(sect)
    chart = PieChart() 
    labels = Reference(pie_sheet, min_col = 1, 
                       min_row = 2, max_row = 900) 
    data = Reference(pie_sheet, min_col = 2, 
                       min_row = 2, max_row = 900)  #When data is delted in B4Files, write is shifted down
    chart.add_data(data, titles_from_data = True)  
    chart.set_categories(labels) 
    chart.title = " PIE-CHART "
    pie_sheet.add_chart(chart, "E2")
    #*******

missing_fields = []
transID_list = []

def missingFields(row, column, exposure_sheet):
    if str(exposure_sheet.iloc[row, column]) == 'nan':
        #**** #https://xlsxwriter.readthedocs.io/working_with_cell_notation.html
        cell = xl_rowcol_to_cell(row, column)
        #****
        transID = exposure_sheet.iloc[row, 0]
        missing_fields.append(cell)
        transID_list.append(transID)
        


#Creating the excel files to output the reports
#***** Code adapted from https://bit.ly/2qopoU6
class reports(tk.Tk):
    def __init__(self, UserID):
        self.__UserID = None
        tk.Tk.__init__(self)
        self.label = tk.Label(self, text='Enter ID', font='Calibri', fg='#003B70', bg='white', width=30)
        self.entry = tk.Entry(self, font='Calibri', fg='#003B70', bg='white', width=30)
        self.button = tk.Button(self, text='Enter',font='Calibri', bg='#003B70', fg='white', width=30,command=self.get_ID)
        self.button.grid(row=2, column=0)
        self.entry.grid(row=1, column=0)
        self.label.grid(row=0, column=0)
#*****

    def get_ID(self):
        re_ID = re.compile('^([a-zA-Z]\d{3}|\d[a-zA-Z]\d{2}|\d{2}[a-zA-Z]\d|\d{3}[a-zA-Z])$')#**** Validation expression from https://bit.ly/2mjCRgR #*****
        new_ID = self.entry.get()
        full_match = re_ID.match(new_ID)
        if full_match != None:
            self.__UserID = self.entry.get()
            messagebox.showinfo('','ID accepted')
        else:
            messagebox.showerror('Error','Please ensure this ID only consists of digits and 1 letter.')

        

    def make_reports(self):
        CQS_workbook = xlsxwriter.Workbook(self.__UserID+'CQS Transaction.xlsx')
        CQS_workbook.close()
        CQSN_workbook = xlsxwriter.Workbook(self.__UserID+'CQS Netting Set.xlsx')
        CQSN_workbook.close()
        LP_workbook = xlsxwriter.Workbook(self.__UserID+'Liquidation Period.xlsx')
        LP_workbook.close()
        SVA_workbook = xlsxwriter.Workbook(self.__UserID+'Security Volatility Adjustment.xlsx')
        SVA_workbook.close()
        FXVA_workbook = xlsxwriter.Workbook(self.__UserID+'FX Volatility Adjustment.xlsx')
        FXVA_workbook.close()
        EV_workbook = xlsxwriter.Workbook(self.__UserID+'Exposure Value.xlsx')
        EV_workbook.close()
        CV_workbook = xlsxwriter.Workbook(self.__UserID+'Collateral Value.xlsx')
        CV_workbook.close()
        NEV_workbook = xlsxwriter.Workbook(self.__UserID+'Net Exposure Value.xlsx')
        NEV_workbook.close()
        BDF_workbook = xlsxwriter.Workbook(self.__UserID+'Blank Data Fields.xlsx')
        BDF_workbook.close()
        CRI_workbook = xlsxwriter.Workbook(self.__UserID+'Credit Rating Inconsistencies.xlsx')
        CRI_workbook.close()

    
    def open_reports(self):
        root = Tk()
        root.title('Select a report')
        root.config()

        CQSTButton = SFTButton(root, text='CQS Transaction',command = lambda:open_report(self.__UserID+'CQS Transaction.xlsx'))
        CQSTButton.grid(row=1)
        CQSNButton = SFTButton(root, text='CQS Netting Set',command = lambda:open_report(self.__UserID+'CQS Netting Set.xlsx'))
        CQSNButton.grid(row=2)
        LPButton = SFTButton(root, text='Liquidation Period',command = lambda:open_report(self.__UserID+'Liquidation Period.xlsx'))
        LPButton.grid(row=3)
        SVAButton = SFTButton(root, text='Security Volatility Adjustment',command = lambda:open_report(self.__UserID+'Security Volatility Adjustment.xlsx'))
        SVAButton.grid(row=4)
        FXVAButton = SFTButton(root, text='FX Volatility Adjustment',command = lambda:open_report(self.__UserID+'FX Volatility Adjustment.xlsx'))
        FXVAButton.grid(row=5)
        EVButton = SFTButton(root, text='Exposure Value',command = lambda:open_report(self.__UserID+'Exposure Value.xlsx'))
        EVButton.grid(row=6)
        CVButton = SFTButton(root, text='Collateral Value',command = lambda:open_report(self.__UserID+'Collateral Value.xlsx'))
        CVButton.grid(row=7)
        NEVButton = SFTButton(root, text='Net Exposure Value',command = lambda:open_report(self.__UserID+'Net Exposure Value.xlsx'))
        NEVButton.grid(row=8)
        BDFButton = SFTButton(root, text='Blank Data Fields',command = lambda:open_report(self.__UserID+'Blank Data Fields.xlsx'))
        BDFButton.grid(row=9)
        CRIButton = SFTButton(root, text='Credit Rating Inconsistencies',command = lambda:open_report(self.__UserID+'Credit Rating Inconsistencies.xlsx'))
        CRIButton.grid(row=10)
                               
        root.mainloop()
    
'''def run_reports():
    create_reports = reports('')
    create_reports.mainloop()
    create_reports.make_reports()
    create_reports.open_reports()'''
    
def run_program(rows_exposure, exposure_sheet, exposure_write, optima_sheet, EE_sheet, columns_exposure, calc_workbook, pie_sheet, lists_sheet, rows_lists, volatility_sheet, rows_volatility):
    create_reports = reports('')
    create_reports.mainloop()
    create_reports.make_reports()
    for row in range(rows_exposure):
        try:
            exposure_write['I'+str(row+2)] = final_transaction(row, exposure_sheet) #row starts at 0 in python and 1 in excel,+2 to keep headers
        except:
            exposure_write['I'+str(row+2)] = 'Error' #except to catch insufficient data and allow the program to keep running
    for row in range(rows_exposure):
        try:
            exposure_write['N'+str(row+2)] = collateral_type(row, optima_sheet)
        except:
            exposure_write['N'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['O'+str(row+2)] = cash_security(row, exposure_sheet)
        except:
            exposure_write['O'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['P'+str(row+2)] = master_netting(row, exposure_sheet)
        except:
            exposure_write['P'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['U'+str(row+2)] = legal_enforceability(row, exposure_sheet)
            exposure_write['V'+str(row+2)] = legal_enforceability(row, exposure_sheet)
            exposure_write['Y'+str(row+2)] = legal_enforceability(row, exposure_sheet)
            exposure_write['Z'+str(row+2)] = legal_enforceability(row, exposure_sheet)
        except:
            exposure_write['U'+str(row+2)] = 'Error'
            exposure_write['V'+str(row+2)] = 'Error'
            exposure_write['Y'+str(row+2)] = 'Error'
            exposure_write['Z'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['W'+str(row+2)] = liquid_illiquid(row, EE_sheet)
        except:
            exposure_write['W'+str(row+2)] = 'Error'
    #for row in range(rows_exposure):
        #print(sufficiently_liquid(row, exposure_sheet))
    for row in range(rows_exposure):
        try:
            exposure_write['AA'+str(row+2)] = correlated_collateral(row, EE_sheet)
        except:
            exposure_write['AA'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['AB'+str(row+2)] = correlation_acceptability(row, EE_sheet)
        except:
            exposure_write['AB'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['AC'+str(row+2)] = SWWR_flag(row, optima_sheet)
        except:
            exposure_write['AC'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['AD'+str(row+2)] = GFP_ID(row, exposure_sheet)
        except:
            exposure_write['AD'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['AF'+str(row+2)] = haircut_security(row, exposure_sheet, optima_sheet)
        except:
            exposure_write['AF'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['AT'+str(row+2)] = basel_asset(row, exposure_sheet, optima_sheet)
        except:
            exposure_write['AT'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['BE'+str(row+2)] = currency_exposure(row, optima_sheet)
        except:
            exposure_write['BE'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['BT'+str(row+2)] = value_collateral(row, exposure_sheet)
        except:
            exposure_write['BT'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['BU'+str(row+2)] = general_requirements(row, exposure_sheet)
        except:
            exposure_write['BU'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['BO'+str(row+2)] = credit_maturity(row, exposure_sheet)
        except:
            exposure_write['BO'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['BX'+str(row+2)] = long_ratingsSP(row, exposure_sheet)
        except:
            exposure_write['BX'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['BY'+str(row+2)] = long_ratingsM(row, exposure_sheet)
        except:
            exposure_write['BY'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['BZ'+str(row+2)] = long_ratingsF(row, exposure_sheet)
        except:
            exposure_write['BZ'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['CA'+str(row+2)] = long_ratingsDBRS(row, exposure_sheet)
        except:
            exposure_write['CA'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['CD'+str(row+2)] = short_ratingsSP(row, exposure_sheet)
        except:
            exposure_write['CD'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['CE'+str(row+2)] = short_ratingsM(row, exposure_sheet)
        except:
            exposure_write['CE'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['CF'+str(row+2)] = short_ratingsF(row, exposure_sheet)
        except:
            exposure_write['CF'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['CG'+str(row+2)] = short_ratingsDBRS(row, exposure_sheet)
        except:
            exposure_write['CG'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['CC'+str(row+2)] = longTermCQS(row, exposure_sheet)
        except:
            exposure_write['CC'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['BW'+str(row+2)] = eligible_collateral(row, exposure_sheet, lists_sheet, rows_exposure)
        except:
            exposure_write['BW'+str(row+2)] = 'Error'
    long_ratingCount(exposure_sheet, rows_exposure)
    for row in range(rows_exposure):
        try:
            exposure_write['CB'+str(row+2)] = long_countif[row]
        except:
            exposure_write['CB'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['BV'+str(row+2)] = collateral_cat(row, exposure_sheet)
        except:
            exposure_write['BV'+str(row+2)] = 'Error'
    short_ratingCount(exposure_sheet, rows_exposure)
    for row in range(rows_exposure):
        try:
            exposure_write['CH'+str(row+2)] = short_countif[row]
        except:
            exposure_write['CH'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['CI'+str(row+2)] = longTermCQS(row, exposure_sheet)
        except:
            exposure_write['CI'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['CK'+str(row+2)] = recognised_exchange(row, exposure_sheet, lists_sheet)
        except:
            exposure_write['CK'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['CM'+str(row+2)] = basic_liquidation(row, exposure_sheet)
        except:
            exposure_write['CM'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['CO'+str(row+2)] = num_trades(row, exposure_sheet)
        except:
            exposure_write['CO'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['CP'+str(row+2)] = large_netting(row, exposure_sheet)
        except:
            exposure_write['CP'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['CR'+str(row+2)] = adjusted_margin(row, exposure_sheet)
        except:
            exposure_write['CR'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['CS'+str(row+2)] = residual_maturity(row, exposure_sheet)
        except:
            exposure_write['CS'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['CT'+str(row+2)] = vol_adj(row, exposure_sheet)
        except:
            exposure_write['CT'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['CU'+str(row+2)] = ten_voladj(row, exposure_sheet, volatility_sheet, rows_volatility)
        except:
            exposure_write['CU'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['CV'+str(row+2)] = security_volatility(row, exposure_sheet)
        except:
            exposure_write['CV'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['CW'+str(row+2)] = security_exposure(row, exposure_sheet)
        except:
            exposure_write['CW'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['CX'+str(row+2)] = fx_mismatch(row, exposure_sheet)
        except:
            exposure_write['CX'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['CY'+str(row+2)] = tenday_fx(row, exposure_sheet)
        except:
            exposure_write['CY'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['CZ'+str(row+2)] = security_fx(row, exposure_sheet)
        except:
            exposure_write['CZ'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['DH'+str(row+2)] = sov_SP(row, exposure_sheet)
        except:
            exposure_write['DH'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['DI'+str(row+2)] = sov_M(row, exposure_sheet)
        except:
            exposure_write['DI'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['DJ'+str(row+2)] = sov_F(row, exposure_sheet)
        except:
            exposure_write['DJ'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['DK'+str(row+2)] = sov_DBRS(row, exposure_sheet)
        except:
            exposure_write['DK'+str(row+2)] = 'Error'
    sov_ratingCount(exposure_sheet, rows_exposure)
    for row in range(rows_exposure):
        try:
            exposure_write['DL'+str(row+2)] = sov_countif[row]
        except:
            exposure_write['DL'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['DM'+str(row+2)] = sov_CQS(row, exposure_sheet)
        except:
            exposure_write['DM'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['DR'+str(row+2)] = sovF_SP(row, exposure_sheet)
        except:
            exposure_write['DR'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['DS'+str(row+2)] = sov_M(row, exposure_sheet)
        except:
            exposure_write['DS'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['DT'+str(row+2)] = sov_F(row, exposure_sheet)
        except:
            exposure_write['DT'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['DU'+str(row+2)] = sov_DBRS(row, exposure_sheet)
        except:
            exposure_write['DU'+str(row+2)] = 'Error'
    sovF_ratingCount(exposure_sheet, rows_exposure)
    for row in range(rows_exposure):
        try:
            exposure_write['DV'+str(row+2)] = sovF_countif[row]
        except:
            exposure_write['DV'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['DW'+str(row+2)] = sovF_CQS(row, exposure_sheet)
        except:
            exposure_write['DW'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['DX'+str(row+2)] = domestic_currency(row, exposure_sheet)
        except:
            exposure_write['DX'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['DY'+str(row+2)] = sovereign_ref(row, exposure_sheet)
        except:
            exposure_write['DY'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['DZ'+str(row+2)] = sov_volA(row, exposure_sheet, volatility_sheet, rows_volatility)
        except:
            exposure_write['DZ'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        try:
            exposure_write['EA'+str(row+2)] = sov_security(row, exposure_sheet)
        except:
            exposure_write['EA'+str(row+2)] = 'Error'
    for row in range(rows_exposure):
        pie_figures(row, exposure_sheet)
    pie_chart(pie_sheet)
    for file in glob.glob('*Blank Data Fields.xlsx'):
        missingFields_workbook = openpyxl.load_workbook(file)
    missingFields_write = missingFields_workbook['Sheet1']
    for column in range(columns_exposure): #Running missing fields function
        for row in range(rows_exposure):
            missingFields(row, column, exposure_sheet)
    for length in range(len(missing_fields)): #Output missing fields function
            missingFields_write['B'+str(length+1)] = missing_fields[length]
            missingFields_write['A'+str(length+1)] = transID_list[length]
    missingFields_workbook.save(file)
    missingFields_workbook.close()
    create_reports.open_reports()

ratings_dict = {
('NR', 'Non rated'):0,
('AAA', 'AA+','AA','AA-', 'Aaa', 'Aa1', 'AA (high)', 'Aa2', 'Aa3', 'AA (low)', 'A-1+','(P)P-1',	'F1+', 'R-1 (high)'):1,
('A+', 'A', 'A-', 'A1', 'A (high)', 'A2', 'A3', 'A (low)','A-1','P-2','F1','R-1 (middle)') :2,
('BBB+', 'BBB','BBB-', 'Baa1', 'BBB (high)', 'Baa2', 'Baa3', 'BBB (low)','A-2','F2','R-1 (low)','A-3', 'P-3', 'F3','R-2 (high)') :3,
('BB+', 'BB', 'BB-', 'Ba1', 'BB (high)', 'Ba2', 'Ba3', 'BB (low)','B','NP','B','R-2 (middle)',	'C' ,'R-2 (low)','R','RD', 'R-3','D' ,'R-4', 'R-5') :4,
('B+', 'B1', 'B (high)', 'B', 'B2', 'B-', 'B3', 'B (low)') :5,
('CCC+', 'Caa1', 'CCC (high)', 'CCC', 'Caa2', 'CCC-', 'Caa', 'C', 'CCC (low)', 'CC', 'Ca', 'RD', 'CC (high)', 'D', 'CC', 'R', 'SD', 'CC (low)', 'C (high)', 'C (low)', 'D') :6
}

basic_liquid = {
('B','BVC','BVDV','BVP','LTRR','LVC','LVDV','LVP','REG','REGR','REV','REVR','S'):5,
('LONG', 'SHORT'):10
}

#***** Code used from https://bit.ly/2KEsEnQ 
working_ratings = {}
for k, v in ratings_dict.items():
    for key in k:
        working_ratings[key] = v
working_basic = {}
for k, v in basic_liquid.items():
    for key in k:
        working_basic[key] = v
#*****  Used to convert ratings_dict into a dictionary
